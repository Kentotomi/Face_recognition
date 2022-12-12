''''
Real Time Face Recogition
	==> Each face stored on dataset/ dir, should have a unique numeric integer ID as 1, 2, 3, etc                       
	==> LBPH computed model (trained faces) should be on trainer/ dir
Based on original code by Anirban Kar: https://github.com/thecodacus/Face-Recognition    

Developed by Marcelo Rovai - MJRoBot.org @ 21Feb18  

'''

import cv2
import numpy as np
import os 
import datetime
import openpyxl
import pyrealsense2 as rs
from PIL import Image

WIDTH = 640
HEIGHT = 480
THRESHOLD = 0.7
SCREEN = 1.7
TARGET = 0.1
list_confidence = [0, 1, 2, 3, 4]


recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.read('trainer/trainer.yml')
cascadePath = "haarcascade_frontalface_default.xml"
faceCascade = cv2.CascadeClassifier(cascadePath);

font = cv2.FONT_HERSHEY_SIMPLEX

#iniciate id counter
id = 0

# names related to ids: example ==> Marcelo: id=1,  etc
names = ['Kento', 'Daiki', 'mask', 'Ilza', 'Z', 'W'] 

# Initialize and start realtime video capture
# color format
# データ形式の話
color_stream, color_format = rs.stream.color, rs.format.bgr8
depth_stream, depth_format = rs.stream.depth, rs.format.z16

# ストリーミング初期化
# RealSenseからデータを受信するための準備
# config.enable_streamでRGB，Dの解像度とデータ形式，フレームレートを指定している
config = rs.config()
config.enable_stream(depth_stream, WIDTH, HEIGHT, depth_format, 30)
config.enable_stream(color_stream, WIDTH, HEIGHT, color_format, 30)

# ストリーミング開始
pipeline = rs.pipeline()
profile = pipeline.start(config)

# 1距離[m] = depth * depth_scale
depth_sensor = profile.get_device().first_depth_sensor()
depth_scale = depth_sensor.get_depth_scale()
# clipping_distance_in_meters = 0.4 # 40cm以内を検出
# depth値にする

# Alignオブジェクト生成
# RGBとDの画角の違いによるズレを修正している
align_to = rs.stream.color
align = rs.align(align_to)
# 検出とプリントするための閾値
# threshold = (WIDTH * HEIGHT * 3) * 0.9
max_dist = THRESHOLD / depth_scale


# cam = cv2.VideoCapture(0)
# cam.set(3, 640) # set video widht
# cam.set(4, 480) # set video height

# Define min window size to be recognized as a face
# minW = 0.1*cam.get(3)
# minH = 0.1*cam.get(4)

dt_now = datetime.datetime.now()

wb = openpyxl.load_workbook('./fr/fr1.xlsx')
ws = wb['Sheet1']
i = 0
try:
    while True:

        # ret, img =cam.read()

        #-------------------realsenseの準備------------------------------
        # フレーム待ち（color&depth）
        # フレーム取得
        frames = pipeline.wait_for_frames()
        # フレームの画角差を修正
        aligned_frames = align.process(frames)
        # フレームの切り分け
        # 多分これに射影変換行列をかけたら視点の変更ができる
        color_frame = aligned_frames.get_color_frame()
        depth_frame = aligned_frames.get_depth_frame()
        if not depth_frame or not color_frame:
            continue



        # dist = depth_frame.get_distance(x, y)

        # RGB画像のフレームから画素値をnumpy配列に変換
        # これで普通のRGB画像になる
        color_image = np.asanyarray(color_frame.get_data())
        

        # D画像のフレームから画素値をnumpy配列に変換
        depth_image = np.asanyarray(depth_frame.get_data()) # 深度の画素値が入っている

        # スクリーンの距離を取得するために3点取得



        # 指定距離以下を無視した深度画像の生成
        # 最大値より遠いものには情報を付与する的な？
        depth_filterd_image = (depth_image < max_dist) * depth_image
        depth_gray_filterd_image = (depth_filterd_image * 255. /max_dist).reshape((HEIGHT, WIDTH)).astype(np.uint8)

        # 指定距離以下を無視したRGB画像の生成
        color_filterd_image = (depth_filterd_image.reshape((HEIGHT, WIDTH, 1)) > 0) * color_image

        # # coverage = [0]*64
        # for y in range(HEIGHT):
        #     for x in range(WIDTH):
        #         dist = depth_frame.get_distance(x, y)
        #         if THRESHOLD < dist and dist < SCREEN - TARGET + 0.05: # 閾値以上スクリーン以下であれば
        #         # リストにその座標を格納するかその画素を消してしまうか
        #             color_filterd_image[y, x] = [0, 255, 0]
        #         #     coverage[x//10] += 1
        img = color_filterd_image



        # img = cv2.flip(img, -1) # Flip vertically

        gray = cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)

        faces = faceCascade.detectMultiScale( 
            gray,
            scaleFactor = 1.2,
            minNeighbors = 5,
            # minSize = (int(minW), int(minH)),
            minSize = (0, 0),
        )

        for(x,y,w,h) in faces:

            cv2.rectangle(img, (x,y), (x+w,y+h), (0,255,0), 2)

            id, confidence = recognizer.predict(gray[y:y+h,x:x+w])



            # Check if confidence is less them 100 ==> "0" is perfect match 
            if (confidence < 100):
                id = names[id]
                confidence = "  {0}%".format(round(100 - confidence))
            else:
                id = "unknown"
                confidence = "  {0}%".format(round(100 - confidence))
            
            
            # yes no 判定処理
            confidence_int = int(confidence[-3:-1])
            if confidence_int < 60:
                confidence_int = 0
            list_confidence[i%5] = confidence_int

            if all(list_confidence):
                print("yes")

            cv2.putText(img, str(id), (x+5,y-5), font, 1, (255,255,255), 2)
            cv2.putText(img, str(confidence), (x+5,y+h-5), font, 1, (255,255,0), 1)
            
            
            # print(dt_now)
            # -----------excel-----------------
            # ws.cell(i,1,value = dt_now)
            # ws.cell(i,2,value = id)
            # wb.save('./fr/fr1.xlsx')
            i = i + 1
            
            
            
        cv2.imshow('camera',img) 

        k = cv2.waitKey(10) & 0xff # Press 'ESC' for exiting video
        if k == 27:
            break
finally:
    pipeline.stop()
    cv2.destroyAllWindows()

# Do a bit of cleanup
# print("\n [INFO] Exiting Program and cleanup stuff")


